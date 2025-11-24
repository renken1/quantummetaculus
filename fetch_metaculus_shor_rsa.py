#!/usr/bin/env python
import datetime as dt
import pathlib

import requests
from openpyxl import Workbook, load_workbook


QUESTION_ID = 3684
API_URL = f"https://www.metaculus.com/api2/questions/{QUESTION_ID}/"
DATA_PATH = pathlib.Path("data/metaculus_shor_rsa_median.xlsx")


def fetch_question_json():
    resp = requests.get(API_URL, timeout=30)
    resp.raise_for_status()
    return resp.json()


def extract_median_date(q_json) -> dt.date:
    """
    Holt das Median-Datum aus einer Metaculus-Date-Frage.

    - continuous_range: Liste von ISO-Zeitstempeln (str)
    - forecast_values: CDF-Werte (0..1) gleicher Länge
    """
    q = q_json["question"]
    scaling = q["scaling"]
    continuous_range = scaling["continuous_range"]

    agg = q["aggregations"]["recency_weighted"]["latest"]
    cdf = agg["forecast_values"]

    if len(continuous_range) != len(cdf):
        raise ValueError("continuous_range and forecast_values length mismatch")

    target = 0.5
    best_idx = min(range(len(cdf)), key=lambda i: abs(cdf[i] - target))

    iso_ts = continuous_range[best_idx]
    # ISO-Strings kommen mit 'Z' -> auf aware datetime mappen, dann date() nehmen
    dt_obj = dt.datetime.fromisoformat(iso_ts.replace("Z", "+00:00"))
    return dt_obj.date()


def ensure_workbook(path: pathlib.Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "metaculus_shor_rsa_median"
        ws.append(["snapshot_date", "median_event_date"])  # Header
    return wb, ws


def append_row(ws, snapshot_date: dt.date, median_date: dt.date):
    """
    Neue Zeile anhängen, falls snapshot_date noch nicht vorhanden ist.
    Wenn die letzte Zeile schon heutiges snapshot_date hat, wird der Wert überschrieben.
    """
    max_row = ws.max_row

    if max_row >= 2:
        last_snapshot = ws.cell(row=max_row, column=1).value
        if isinstance(last_snapshot, dt.datetime):
            last_snapshot = last_snapshot.date()
        if last_snapshot == snapshot_date:
            ws.cell(row=max_row, column=2).value = median_date
            return

    ws.append([snapshot_date, median_date])


def main():
    snapshot_date = dt.date.today()
    q_json = fetch_question_json()
    median_date = extract_median_date(q_json)

    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb, ws = ensure_workbook(DATA_PATH)
    append_row(ws, snapshot_date, median_date)
    wb.save(DATA_PATH)


if __name__ == "__main__":
    main()
